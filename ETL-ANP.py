#!/usr/bin/env python
# coding: utf-8

# In[4]:


import os
from urllib                import request
import glob
import pandas              as pd
from openpyxl              import load_workbook
import numpy               as np
import datetime
import pyarrow.parquet     as pq
import pyarrow             as pa  
import datetime            as dt


# In[11]:


#Definição das variaveis
vURL                 = 'https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls'
vCaminho             = './anp/raw/'
vArquivo             = './anp/raw/vendas-combustiveis-m3.xls'
vArquivo_xlsx        = './anp/raw/vendas-combustiveis-m3.xlsx'
vBaseStaging         =  './base_staging.xlsx'
vCaminhoStaging      = './*.xlsx'
vFormatoData         = '%Y/%m/%d'
vParticao            = str(dt.datetime.now().strftime('%Y/%m/%d'))
vLocationParquet     = './anp/raw/parquet/'
vCompressao          = "snappy"


# In[3]:


#Criar a pasta caso não exista 
isExist = os.path.exists(vCaminho)
if not isExist:  
    os.makedirs(vCaminho) 
    print('Pasta foi criada')
#Download do arquivo
request.urlretrieve(vURL , vArquivo )
#Conversão do arquivo para leitura das Planilhas ocultas
os.system(f'''libreoffice --headless --invisible --convert-to xlsx {vArquivo} --outdir {vCaminho}''')


# In[4]:


#Check do download
x = [os.path.join(r,file) for r,d,f in os.walk(vCaminho) for file in f]
print(x)


# In[5]:


#Identificar quais arquivos devem ser gerados a partir da Planilha principal
workBook                    = load_workbook(vArquivo_xlsx)
sheets                      = workBook.sheetnames
list_carga                  = sheets
list_carga.pop(0)
print(list_carga)


# In[6]:


#Função para fazer o split do arquivo com as tabelas dinamicas
def split_xlsx(aba_excel,arquivo,destino_arquivo):
    if os.path.exists(arquivo):
        os.remove(arquivo)
    workBook = load_workbook(destino_arquivo)
    sheets = workBook.sheetnames
    for sheet in sheets:
        if sheet != aba_excel:
            workBook.remove(workBook.get_sheet_by_name(sheet))    
    workBook[aba_excel].sheet_state = 'visible'
    workBook.save(arquivo)


# In[7]:


#Exportação das dinamicas
for i in list_carga:
    split_xlsx(i,f'base_staging_{i}.xlsx',vArquivo_xlsx)  
    print('Finalizou a extração do arquivo ', i)


# In[8]:


#Lista de arqruivos contidos na pasta
vValida_arquivos          = glob.glob(vCaminhoStaging)
print(vValida_arquivos)


# In[9]:


#Estrutura que compoe  a base bruta para ingestão


# In[14]:


print('Carga Iniciada ' + str(dt.datetime.now().strftime('%d/%m/%Y %H:%M:%S')))

##Leitura e concatenação dos dataframes
df_m3                     = pd.read_excel('./base_staging_DPCache_m3.xlsx')
df_m3_2                   = pd.read_excel('./base_staging_DPCache_m3_2.xlsx')
df_bruto                  = pd.concat([df_m3, df_m3_2])

#Foi preciso reorganizar as columas após a exportação dos arquivos bases a partir do xls desconfigurou as colunas
df_col= ['product', 'year', 'region', 'uf','total', 'Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
df_bruto.columns=df_col

##Transformação do dataframe
df_bruto                  = df_bruto.drop(columns=['total'])
df_bruto.columns          = df_bruto.columns.str.lower()
df_bruto['unit']          = df_bruto["product"].apply(lambda x: x[-3:-1])
df_bruto["product"]       = df_bruto["product"].apply(lambda x: x[0:-4].strip())
df_unpivot                = df_bruto.melt(id_vars=['product', 'unit', 'region', 'uf','year'],
                                          var_name='month',  value_name  = 'volume')
df_unpivot['volume']      = df_unpivot['volume'].replace(np.nan, 0)
df_unpivot['volume']      = df_unpivot['volume'].replace(np.nan, 0)

##Foi necessario converter o mês para executar a transformação da data 
months                    ={'jan':'01', 'fev':'02', 'mar':'03', 'abr':'04', 'mai':'05', 'jun':'06',
                            'jul':'07', 'ago':'08', 'set':'09', 'out':'10', 'nov':'11', 'dez':'12'}
df_unpivot['month']       = df_unpivot.month.replace(months)
df_unpivot["year_month"]  = df_unpivot['year'].astype(str) +"-"+ df_unpivot["month"]

#Remoção de colunas desnecessarias
df_unpivot                = df_unpivot.drop(columns=['region','year', 'month'])

#Manipulação de datas
df_unpivot['year_month']  = pd.to_datetime(df_unpivot['year_month'], errors = "coerce")
df_unpivot['created_at']  = datetime.datetime.now().strftime(vFormatoData)
df_unpivot['created_at']  = df_unpivot['created_at'].astype('datetime64[ns]')

#Reorganização das colunas para o Dataframe
df_unpivot                = df_unpivot[["year_month", "uf", "product","unit", "volume", "created_at"]]
df_unpivot                = df_unpivot.groupby(['year_month','uf','product','unit', "created_at"])[['volume']].sum().reset_index()

#Criação da tabela transformando em Parquet
vTableFileName             = 'base_derivados'
df_derivados               = df_unpivot[~df_unpivot['product'].str.contains('ÓLEO', na=False)]
vTable2Extract_derivados   = pa.Table.from_pandas(df_derivados,preserve_index = False)
pa.parquet.write_to_dataset(table                 = vTable2Extract_derivados,
                            root_path             = vLocationParquet, 
                            partition_filename_cb = lambda x: vTableFileName + ".parquet",
                            compression           = vCompressao)


vTableFileName            = 'base_oleo'
df_oleo                   = df_unpivot[df_unpivot['product'].str.contains('ÓLEO', na=False)]
vTable2Extract_oleo       = pa.Table.from_pandas(df_oleo,preserve_index = False)
pa.parquet.write_to_dataset(table                 = vTable2Extract_oleo,
                            root_path             = vLocationParquet, 
                            partition_filename_cb = lambda x: vTableFileName + ".parquet",
                            compression           = vCompressao)
print('Carga finalizada ' + str(dt.datetime.now().strftime('%d/%m/%Y %H:%M:%S')))

